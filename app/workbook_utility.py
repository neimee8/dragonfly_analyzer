from pathlib import Path
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

class WorkbookUtility:
    def __init__(self, result_file):
        self.result_file = result_file
        self.wb = self._set_or_init_workbook(result_file)
        self.current_column = 1

    def add_sheet(self, sheet_name):
        self.current_column = 1
        return self.wb.create_sheet(title=sheet_name)

    def call_column(self):
        old = self.current_column
        self.current_column += 1
        return old

    def _set_or_init_workbook(self, result_file):
        if Path(result_file).exists():
            return load_workbook(result_file)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                std = wb["Sheet"]
                wb.remove(std)
            return wb

    def save(self):
        self.wb.save(self.result_file)
